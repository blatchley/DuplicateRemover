import openpyxl as pyxl
from datetime import datetime
import configparser
import os

## A short python program designed to filter out duplicate test IDs from very large excel datasets.
## Uses pyxl read and write only workbooks to enable very large data sets to be filtered efficiently with no memory overhead.
## -Adam B. Hansen

## Get config.ini
__location__ = os.path.realpath(os.path.join(os.getcwd(), os.path.dirname(__file__)))
Config = configparser.ConfigParser()
Config.read(os.path.join(__location__, 'config.ini'))


## Load information from config.ini
inputFileName = Config['Path']['InputFileName']
outputFileName = Config['Path']['OutputFileName']
sheetToLoad = int(Config['Path']['SheetIndex'])


## Create read only pyxl workbook using the input file
print("preparing to load workbook " + inputFileName)
inputBook = pyxl.load_workbook(filename=inputFileName, read_only=True)
print("workbook loaded")

## Load correct sheet
print("loading sheet " + inputBook.sheetnames[sheetToLoad])
ws = inputBook[inputBook.sheetnames[sheetToLoad]]

## Create write only pyxl workbook to write sorted data to.
filteredBook = pyxl.Workbook(write_only=True)
wsF = filteredBook.create_sheet()


print("filtering...")



## Iterates over list of entries once
previousId = 0
currentBestRow = []
for row in ws.rows:
	## If row shares same ID as previous row, save the one with the earliest date
	if(row[0].value == previousId):
		rowContents = []
		for cell in row:
			if(type(cell.value) == datetime):
				rowContents.append(cell.value.strftime("%d/%m/%Y %H:%M"))
			else:
				rowContents.append(cell.value)
		dateThis = datetime.strptime(rowContents[4],"%d/%m/%Y %H:%M")
		dateBest = datetime.strptime(currentBestRow[4],"%d/%m/%Y %H:%M")
		## Keep row version with earliest date
		if(dateThis < dateBest):
			currentBestRow = rowContents
		previousId = rowContents[0]
	## If row does not share ID of previous row
	else:
		## Add earliest blood test with previous ID to output file
		wsF.append(currentBestRow)
		## Create first element for new ID.
		rowContents = []
		for cell in row:
			if(type(cell.value) == datetime):
				rowContents.append(cell.value.strftime("%d/%m/%Y %H:%M"))
			else:
				rowContents.append(cell.value)
		currentBestRow = rowContents
		previousId = rowContents[0]

## On end of parsing, append earliest row for final ID
wsF.append(currentBestRow)		

print("Saving Filtered worksheet to file " +outputFileName)


## Save output workbook
filteredBook.save(outputFileName) 
